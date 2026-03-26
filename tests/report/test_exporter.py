from pathlib import Path

from openpyxl import load_workbook

from app.report.analyzer import AnalysisResult, LemmaStats
from app.report.excel_exporter import ExcelExporter


def test_export_creates_xlsx_file():
    exporter = ExcelExporter()
    analysis_result = AnalysisResult(
        total_lines=2,
        lemma_stats={
            "дом": LemmaStats(
                total_count=2,
                line_counts={0: 1, 1: 1},
            ),
        },
    )

    output_file_path = exporter.export(analysis_result)

    assert Path(output_file_path).exists()
    assert output_file_path.endswith(".xlsx")


def test_export_writes_header():
    exporter = ExcelExporter()
    analysis_result = AnalysisResult(
        total_lines=3,
        lemma_stats={
            "дом": LemmaStats(
                total_count=2,
                line_counts={0: 1, 2: 1},
            ),
            "кот": LemmaStats(
                total_count=1,
                line_counts={1: 1},
            ),
        },
    )

    output_file_path = exporter.export(analysis_result)

    workbook = load_workbook(output_file_path)
    worksheet = workbook["Отчет"]

    rows = list(worksheet.iter_rows(values_only=True))

    assert rows[0] == (
        "Словоформа",
        "Количество во всём документе",
        "Количество по строкам",
    )


def test_export_sorts_lemmas_alphabetically():
    exporter = ExcelExporter()
    analysis_result = AnalysisResult(
        total_lines=1,
        lemma_stats={
            "яблоко": LemmaStats(total_count=1, line_counts={0: 1}),
            "арбуз": LemmaStats(total_count=1, line_counts={0: 1}),
            "дыня": LemmaStats(total_count=1, line_counts={0: 1}),
            "виноград": LemmaStats(total_count=1, line_counts={0: 1}),
        },
    )

    output_file_path = exporter.export(analysis_result)

    workbook = load_workbook(output_file_path)
    worksheet = workbook["Отчет"]

    rows = list(worksheet.iter_rows(values_only=True))

    assert rows[1][0] == "арбуз"
    assert rows[2][0] == "виноград"
    assert rows[3][0] == "дыня"
    assert rows[4][0] == "яблоко"


def test_fills_missing_lines_with_zero():
    exporter = ExcelExporter()

    result = exporter._line_counts_to_str(
        line_counts={1: 11, 2: 32, 5: 3},
        total_lines=6,
    )

    assert result == "0,11,32,0,0,3"


def test_export_empty_analysis_result():
    exporter = ExcelExporter()
    analysis_result = AnalysisResult(
        total_lines=0,
        lemma_stats={},
    )

    output_file_path = exporter.export(analysis_result)

    workbook = load_workbook(output_file_path)
    worksheet = workbook["Отчет"]

    rows = list(worksheet.iter_rows(values_only=True))

    assert len(rows) == 1
    assert rows[0] == (
        "Словоформа",
        "Количество во всём документе",
        "Количество по строкам",
    )
    

def test_export_writes_rows():
    exporter = ExcelExporter()
    analysis_result = AnalysisResult(
        total_lines=3,
        lemma_stats={
            "дом": LemmaStats(
                total_count=2,
                line_counts={0: 1, 2: 1},
            ),
            "кот": LemmaStats(
                total_count=1,
                line_counts={1: 1},
            ),
        },
    )

    output_file_path = exporter.export(analysis_result)

    workbook = load_workbook(output_file_path)
    worksheet = workbook["Отчет"]

    rows = list(worksheet.iter_rows(values_only=True))

    assert rows[1] == ("дом", 2, "1,0,1")
    assert rows[2] == ("кот", 1, "0,1,0")
