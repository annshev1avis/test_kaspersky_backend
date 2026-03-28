import asyncio
from io import BytesIO
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app
from app.report.router import MAX_FILE_SIZE_MB, MAX_FILE_SIZE_BYTES
import app.report.router as report_router_module

client = TestClient(app)


def read_xlsx_response(content: bytes):
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    return rows
    
    
def test_export_report_returns_xlsx_file():
    file_content = "Житель жил в доме\nЖителем гордились\n"
    response = client.post(
        "/public/report/export",
        files={"file": ("input.txt", file_content.encode("utf-8"), "text/plain")}
    )
    
    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment; filename=\"report.xlsx\"" in response.headers["content-disposition"]
    

def test_export_report_contains_expected_statistics():
    file_content = "Житель жил\nЖителем гордились\n"
    response = client.post(
        "/public/report/export",
        files={"file": ("input.txt", file_content.encode("utf-8"), "text/plain")},
    )

    assert response.status_code == 200

    rows = read_xlsx_response(response.content)

    assert rows[0] == ("Словоформа", "Количество во всём документе", "Количество по строкам")

    data_rows = rows[1:]
    data = {row[0]: row[1:] for row in data_rows}

    assert "житель" in data
    assert data["житель"] == (2, "1,1")


def test_export_report_returns_400_for_empty_file():
    response = client.post(
        "/public/report/export",
        files={"file": ("empty.txt", b"", "text/plain")},
    )

    assert response.status_code == 400
    assert response.json() == {"detail": "Загруженный файл не должен быть пустым"}


def test_export_report_returns_400_for_non_utf8_file():
    file_content_cp1251 = "Привет".encode("cp1251")

    response = client.post(
        "/public/report/export",
        files={"file": ("input.txt", file_content_cp1251, "text/plain")},
    )

    assert response.status_code == 400
    assert response.json() == {"detail": "Должен быть передан текстовый файл в кодировке utf-8"}
    

def test_export_report_returns_413_for_too_large_file():
    oversized_content = b"a" * (MAX_FILE_SIZE_BYTES + 1)

    response = client.post(
        "/public/report/export",
        files={"file": ("large.txt", oversized_content, "text/plain")},
    )

    assert response.status_code == 413
    assert f"Размер файла не должен превышать {MAX_FILE_SIZE_MB} МБ" in response.json()["detail"]


def test_export_report_returns_503_when_server_is_overloaded(monkeypatch):
    monkeypatch.setattr(
        report_router_module,
        "report_semaphore",
        asyncio.Semaphore(0),
    )

    response = client.post(
        "/public/report/export",
        files={"file": ("input.txt", "Привет\n".encode("utf-8"), "text/plain")},
    )

    assert response.status_code == 503
    assert response.json() == {
        "detail": "Сервер временно перегружен. Попробуйте снова через пару минут"
    }